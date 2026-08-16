"""workbook.py — fill the four supplied .xlsx templates. The submission itself.

Without this module there is no entry: OpenStocks accepts workbooks, not JSON.
Everything here exists to make one of the twelve numbers impossible to lose to a
mechanical error, because a mechanical error costs exactly as much as being
wrong about the business.

## The rules this module enforces (SUBMISSION.md)

  1. START FROM THE SUPPLIED TEMPLATE. Never construct a workbook. The template
     is copied byte-for-byte, then three cells are assigned. The `Summary` sheet
     name, the `Instructions` sheet, the metric labels, the units and the fiscal
     period header are never touched — they are what the checker keys on.
  2. FILL ONLY THE YELLOW CELLS. The forecast column is discovered from the
     header row and each row is located by MATCHING ITS LABEL TEXT, never by a
     hardcoded row number, so a template revision surfaces as a loud failure
     instead of a number written into the wrong row.
  3. UNITS. Percentages in POINTS (4.5 = 4.5%). Hays EPS in PENCE. Money in
     MILLIONS. Each is checked against a band and shouted about on breach.
  4. NUMERIC CELLS. `float()`, never `str()`. `check-forecasts.mjs` rejects a
     string, and a string is also unscoreable.
  5. NEVER BLANK. A missing forecast scores the maximum 5.0 penalty — the worst
     outcome available for a metric. Every path through this module writes a
     number. Where that number is a fallback rather than a forecast, it is
     printed in a banner that is impossible to miss in a run log.

## The real template structure (discovered with openpyxl, not assumed)

All four templates share one layout:

    Sheets:  ['Summary', 'Instructions']
    Summary A1:C12
      A6 'Metric'   B6 'Units'   C6 <fiscal period>     <- header, blue FF155EEF
      A7 <label 1>  B7 <units>   C7 forecast            <- yellow FFFFF7D6
      A8 <label 2>  B8 <units>   C8 forecast            <- yellow FFFFF7D6
      A9 <label 3>  B9 <units>   C9 forecast            <- yellow FFFFF7D6

Note that A4 — the "Enter forecasts only in the yellow cells" banner — carries
the SAME yellow fill as the forecast cells. Yellow alone is therefore not a
sufficient locator, which is exactly why the header row and the label text are
the anchors and the fill is only a cross-check.

## Usage

    from submission.workbook import write_workbooks
    paths = write_workbooks(forecasts)              # -> starter/submission/*.xlsx

    python -m submission.workbook --from runs/forecasts
    python -m submission.workbook --from runs/forecasts --dry-run
    python -m submission.workbook --synthetic --out starter/submission
    python -m submission.workbook --verify-only

Then, from inside `starter/`:  npm run check:forecasts
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import os
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl

logger = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = REPO_ROOT / "starter" / "challenge" / "templates"
DEFAULT_OUT_DIR = REPO_ROOT / "starter" / "submission"
DEFAULT_FORECAST_DIR = REPO_ROOT / "runs" / "forecasts"
COMPANIES_JSON = REPO_ROOT / "starter" / "challenge" / "companies.json"

SUMMARY_SHEET = "Summary"

# The fill on the three forecast cells, read off the supplied templates.
# A4 shares it, so this is a cross-check and never a locator.
YELLOW = {"FFFFF7D6", "FFF7D6"}

# `check-forecasts.mjs` reads the forecast from column 3 of the row that follows
# the Metric/Units/<period> header, in companies.json order. Both facts are
# asserted against the template rather than trusted.
VALIDATOR_FORECAST_COLUMN = 3
HEADER_SEARCH_ROWS = 30


# ---------------------------------------------------------------------------
# 1. Reuse from analysts/, with a degraded path that still ships a workbook
# ---------------------------------------------------------------------------

# `analysts.central` pulls in the whole agent stack (agent_core -> openai-agents
# -> litellm). That is the right dependency for the pipeline and the wrong one
# for the last step of the day: if the agent stack fails to import at 17:55 the
# workbooks must still be written. So the canonical definitions are imported
# when they can be, and reconstructed from challenge/companies.json when they
# cannot — loudly, because a divergence between the two is a real risk.

_CENTRAL_IMPORT_ERROR: str = ""

try:  # pragma: no cover - exercised by the degraded path only
    from analysts.central import (  # type: ignore[attr-defined]
        COMPANY_METRICS,
        canonical_ticker,
        metrics_for,
        output_file_for,
        period_for,
        validate_forecast,
    )
    from analysts.central import _MAGNITUDE_BOUNDS as MAGNITUDE_BOUNDS  # noqa: N812
except Exception as _e:  # pragma: no cover
    _CENTRAL_IMPORT_ERROR = f"{type(_e).__name__}: {_e}"
    validate_forecast = None  # type: ignore[assignment]
    MAGNITUDE_BOUNDS: dict[tuple[str, str], tuple[float, float]] = {}

    def canonical_ticker(ticker: str) -> str:  # type: ignore[misc]
        """`LSE:HAS`, `HAS.L`, `has` -> `HAS`. Mirrors analysts.central."""
        t = (ticker or "").strip().upper()
        if ":" in t:
            t = t.rsplit(":", 1)[-1]
        for suffix in (".L", ".US", ".UK"):
            if t.endswith(suffix):
                t = t[: -len(suffix)]
        return t

    def _load_company_metrics() -> dict[str, dict[str, Any]]:
        raw = json.loads(COMPANIES_JSON.read_text(encoding="utf-8"))
        return {
            canonical_ticker(e["ticker"]): {
                "company": e.get("company", ""),
                "ticker": e["ticker"],
                "period": e.get("period", ""),
                "output_file": e.get("outputFile", ""),
                "metrics": [
                    {"label": m["label"], "units": m["units"]} for m in e.get("metrics", [])
                ],
            }
            for e in raw["companies"]
        }

    COMPANY_METRICS = _load_company_metrics()  # type: ignore[misc]

    def metrics_for(ticker: str) -> list[dict[str, str]]:  # type: ignore[misc]
        return list(COMPANY_METRICS.get(canonical_ticker(ticker), {}).get("metrics", []))

    def period_for(ticker: str) -> str:  # type: ignore[misc]
        return COMPANY_METRICS.get(canonical_ticker(ticker), {}).get("period", "")

    def output_file_for(ticker: str) -> str:  # type: ignore[misc]
        return COMPANY_METRICS.get(canonical_ticker(ticker), {}).get("output_file", "")


TICKERS: tuple[str, ...] = tuple(COMPANY_METRICS)


# ---------------------------------------------------------------------------
# 2. Output
# ---------------------------------------------------------------------------


def _stdout_utf8() -> None:
    """A Windows console is cp1252; one em dash from a metric label aborts the
    final step of the day. Presentation must never kill a finished forecast."""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
        except Exception:
            pass


def say(text: str = "") -> None:
    print(text, flush=True)


def shout(lines: Sequence[str]) -> None:
    """A banner for anything a human must not scroll past."""
    width = max(74, max((len(x) for x in lines), default=0) + 6)
    say("!" * width)
    for line in lines:
        say(f"!! {line}")
    say("!" * width)


# ---------------------------------------------------------------------------
# 3. Template layout — discovered, never assumed
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class MetricRow:
    label: str
    units: str
    row: int
    label_cell: str
    units_cell: str
    forecast_cell: str
    number_format: str
    is_yellow: bool


@dataclass
class TemplateLayout:
    ticker: str
    period: str
    sheet: str
    header_row: int
    forecast_column: int
    period_header_cell: str
    rows: list[MetricRow] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)


def _text(cell_value: Any) -> str:
    return "" if cell_value is None else str(cell_value).strip()


def _fill_rgb(cell: Any) -> str:
    fill = cell.fill
    if fill is None or not fill.fill_type:
        return ""
    rgb = getattr(fill.fgColor, "rgb", None)
    return str(rgb).upper() if isinstance(rgb, str) else ""


def read_layout(path: Path, ticker: str, *, is_template: bool = True) -> TemplateLayout:
    """Open a template (or a written workbook) and find the real forecast cells.

    Anchors, in order: the `Summary` sheet, the Metric/Units/<period> header row,
    then each metric's own LABEL TEXT. The row number is an output of this
    function, never an input. The yellow fill and the validator's fixed row/column
    expectations are then cross-checked against what was found, so a revised
    template fails here rather than silently misplacing a number.
    """
    tk = canonical_ticker(ticker)
    period = period_for(tk)
    expected = metrics_for(tk)
    problems: list[str] = []

    wb = openpyxl.load_workbook(path)
    try:
        if SUMMARY_SHEET not in wb.sheetnames:
            problems.append(
                f"ERROR: {path.name}: no {SUMMARY_SHEET!r} sheet (found {wb.sheetnames}). "
                f"check-forecasts.mjs reads workbook.Sheets.Summary and nothing else."
            )
            return TemplateLayout(tk, period, SUMMARY_SHEET, 0, 0, "", [], problems)
        ws = wb[SUMMARY_SHEET]

        # --- the header row ------------------------------------------------
        # Exactly the validator's own search: A == "Metric", B == "Units",
        # C == the fiscal period, within the first 30 rows.
        header_row = 0
        for r in range(1, HEADER_SEARCH_ROWS + 1):
            if (
                _text(ws.cell(r, 1).value) == "Metric"
                and _text(ws.cell(r, 2).value) == "Units"
                and _text(ws.cell(r, 3).value) == period
            ):
                header_row = r
                break
        if not header_row:
            problems.append(
                f"ERROR: {path.name}: no Metric / Units / {period!r} header row in the "
                f"first {HEADER_SEARCH_ROWS} rows. The template's period header does not "
                f"match challenge/companies.json."
            )
            return TemplateLayout(tk, period, SUMMARY_SHEET, 0, 0, "", [], problems)

        forecast_column = VALIDATOR_FORECAST_COLUMN
        period_cell = ws.cell(header_row, forecast_column).coordinate

        # --- each metric row, found by its label ---------------------------
        rows: list[MetricRow] = []
        for index, spec in enumerate(expected):
            label, units = spec["label"], spec["units"]
            found = 0
            for r in range(header_row + 1, header_row + 12):
                if _text(ws.cell(r, 1).value) == label:
                    found = r
                    break
            if not found:
                problems.append(
                    f"ERROR: {path.name}: metric label {label!r} does not appear in "
                    f"column A below the header. The template no longer matches "
                    f"challenge/companies.json — do NOT edit the label, investigate."
                )
                continue

            # The validator does not search: it reads header_row + index + 1.
            # A label that exists but sits elsewhere would still fail the check,
            # so the discovered row has to agree with that.
            required = header_row + index + 1
            if found != required:
                problems.append(
                    f"ERROR: {path.name}: {label!r} is on row {found} but "
                    f"check-forecasts.mjs reads row {required} for metric #{index + 1}. "
                    f"The template's metric order diverges from companies.json."
                )

            unit_cell_text = _text(ws.cell(found, 2).value)
            if unit_cell_text != units:
                problems.append(
                    f"ERROR: {path.name}: {label!r} units cell reads {unit_cell_text!r}, "
                    f"companies.json says {units!r}. The units are NOT ours to change — "
                    f"the checker compares them verbatim."
                )

            target = ws.cell(found, forecast_column)
            is_yellow = _fill_rgb(target) in YELLOW
            if not is_yellow:
                problems.append(
                    f"WARN: {path.name}!{target.coordinate} ({label}) is not the expected "
                    f"yellow {sorted(YELLOW)[0]} fill (found {_fill_rgb(target) or 'none'}). "
                    f"Only yellow forecast cells should be filled — verify by eye."
                )
            if is_template and target.value is not None:
                problems.append(
                    f"WARN: {path.name}!{target.coordinate} ({label}) already holds "
                    f"{target.value!r} in the source template; it will be overwritten."
                )

            rows.append(
                MetricRow(
                    label=label,
                    units=units,
                    row=found,
                    label_cell=ws.cell(found, 1).coordinate,
                    units_cell=ws.cell(found, 2).coordinate,
                    forecast_cell=target.coordinate,
                    number_format=target.number_format,
                    is_yellow=is_yellow,
                )
            )

        return TemplateLayout(
            ticker=tk,
            period=period,
            sheet=SUMMARY_SHEET,
            header_row=header_row,
            forecast_column=forecast_column,
            period_header_cell=period_cell,
            rows=rows,
            problems=problems,
        )
    finally:
        wb.close()


def template_path(ticker: str) -> Path:
    name = output_file_for(canonical_ticker(ticker))
    if not name:
        raise KeyError(
            f"unknown ticker {ticker!r}; challenge/companies.json knows {', '.join(TICKERS)}"
        )
    return TEMPLATE_DIR / name


def dump_templates() -> None:
    """Print the discovered structure of all four supplied templates."""
    for tk in TICKERS:
        p = template_path(tk)
        say("=" * 78)
        say(f"{p.name}   ({tk}, {period_for(tk)})")
        say("=" * 78)
        if not p.exists():
            shout([f"TEMPLATE MISSING: {p}"])
            continue
        wb = openpyxl.load_workbook(p)
        say(f"  sheets           : {wb.sheetnames}")
        wb.close()
        layout = read_layout(p, tk)
        say(f"  header row       : {layout.header_row}  "
            f"(period header at {layout.sheet}!{layout.period_header_cell})")
        say(f"  forecast column  : {layout.forecast_column}")
        for mr in layout.rows:
            say(f"    {mr.forecast_cell:<4} <- {mr.label:<42} [{mr.units:<11}] "
                f"label@{mr.label_cell} units@{mr.units_cell} "
                f"fmt={mr.number_format!r} yellow={mr.is_yellow}")
        for p_ in layout.problems:
            say(f"  {p_}")
        say()


# ---------------------------------------------------------------------------
# 4. Accepting forecasts in whatever shape they arrive
# ---------------------------------------------------------------------------


@dataclass
class SourceMetric:
    label: str = ""
    units: str = ""
    value: float | None = None
    low: float | None = None
    high: float | None = None


@dataclass
class SourceForecast:
    ticker: str = ""
    period: str = ""
    metrics: list[SourceMetric] = field(default_factory=list)
    origin: str = ""  # where it came from, for the log
    raw: Any = None  # the original CompanyForecast, if we were handed one


def _num(x: Any) -> float | None:
    """A finite float, or None. NaN and inf are missing values, not numbers."""
    if x is None or isinstance(x, bool):
        return None
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    return v if math.isfinite(v) else None


def _to_mapping(obj: Any) -> dict[str, Any] | None:
    if isinstance(obj, dict):
        return obj
    dump = getattr(obj, "model_dump", None)  # pydantic v2 (CompanyForecast)
    if callable(dump):
        try:
            return dump(mode="json")
        except TypeError:
            return dump()
    return None


def normalise(obj: Any, *, origin: str = "") -> SourceForecast:
    """`CompanyForecast`, a dict, or the pipeline's JSON envelope -> SourceForecast.

    Deliberately permissive about the container and strict about the contents:
    the numbers are what matter and the labels are re-derived from
    challenge/companies.json regardless, so a wrapper shape can never lose one.
    """
    raw = obj
    data = _to_mapping(obj)
    if data is None:
        return SourceForecast(origin=origin, raw=raw)

    # The pipeline writes {"run_id":..., "workbook":..., "forecast": {...}}.
    if "forecast" in data and isinstance(data["forecast"], dict):
        raw = obj
        data = data["forecast"]

    metrics: list[SourceMetric] = []
    for m in data.get("metrics") or []:
        md = _to_mapping(m) or {}
        metrics.append(
            SourceMetric(
                label=str(md.get("label") or ""),
                units=str(md.get("units") or ""),
                value=_num(md.get("value")),
                low=_num(md.get("low")),
                high=_num(md.get("high")),
            )
        )
    return SourceForecast(
        ticker=canonical_ticker(str(data.get("ticker") or "")),
        period=str(data.get("period") or ""),
        metrics=metrics,
        origin=origin,
        raw=raw,
    )


def _norm_label(text: str) -> str:
    return "".join(ch for ch in (text or "").lower() if ch.isalnum())


def _match_metric(source: SourceForecast, label: str) -> SourceMetric | None:
    key = _norm_label(label)
    for m in source.metrics:
        if m.label == label:
            return m
    for m in source.metrics:
        if _norm_label(m.label) == key:
            return m
    return None


# ---------------------------------------------------------------------------
# 5. Units — the cheapest way to throw a metric, so shout about it
# ---------------------------------------------------------------------------

PERCENT_FRACTION_THRESHOLD = 0.15  # |v| below this in a % cell smells like 0.045
PENCE_POUNDS_THRESHOLD = 0.5       # |v| below this in a GBp cell smells like GBP


def unit_warnings(ticker: str, label: str, units: str, value: float) -> list[str]:
    """Loud, specific complaints about a value that looks mis-scaled.

    These are ORDER-OF-MAGNITUDE checks, not opinions about the forecast. Each
    one names the metric, because a warning that does not say which of twelve
    numbers it is about gets ignored.
    """
    tk = canonical_ticker(ticker)
    u = (units or "").strip()
    out: list[str] = []
    tag = f"{tk} / {label}"

    is_pct = u == "%"
    is_pence = u == "GBp"
    is_money = u in {"USDm", "GBPm"}
    is_eps = "eps" in _norm_label(label) or "share" in u.lower()

    if is_pct:
        if value != 0.0 and abs(value) < PERCENT_FRACTION_THRESHOLD:
            out.append(
                f"UNITS: {tag}: {value} in a PERCENTAGE cell. Percentages are in POINTS, "
                f"so this workbook says {value}%. If you meant {value * 100:.2f}% this is "
                f"wrong by 100x."
            )
        if abs(value) > 100.0:
            out.append(
                f"UNITS: {tag}: {value} exceeds 100 percentage points — this looks like a "
                f"fraction multiplied out, or basis points ({value / 100:.2f}%?)."
            )
    if is_pence:
        if value != 0.0 and abs(value) < PENCE_POUNDS_THRESHOLD:
            out.append(
                f"UNITS: {tag}: {value} in a PENCE (GBp) cell looks like POUNDS. "
                f"{value} pounds is {value * 100:.1f}p. Hays EPS is quoted in pence."
            )
        if abs(value) > 500.0:
            out.append(
                f"UNITS: {tag}: {value}p is implausible for Hays EPS — pence, not a "
                f"pence-per-thousand or a money total."
            )
    if is_money and abs(value) >= 1_000_000:
        out.append(
            f"UNITS: {tag}: {value:,.0f} in a MILLIONS cell is {value / 1_000_000:,.1f} "
            f"trillion. Money is in millions."
        )
    if is_eps and not is_pence and abs(value) > 100.0:
        out.append(f"UNITS: {tag}: EPS of {value} is implausible for a per-share figure.")

    bounds = MAGNITUDE_BOUNDS.get((tk, label))
    if bounds:
        lo, hi = bounds
        if not (lo <= value <= hi):
            hint = ""
            if is_money and value != 0:
                if value < lo and lo / max(abs(value), 1e-9) > 100:
                    hint = " Looks like BILLIONS in a millions cell (x1000)."
                elif value > hi and abs(value) / hi > 100:
                    hint = " Looks like units in a millions cell (/1000)."
            out.append(
                f"MAGNITUDE: {tag}: {value:,g} {u} is outside the sanity band "
                f"[{lo:,g}, {hi:,g}]. That band is an order-of-magnitude guardrail several "
                f"times wider than any plausible outcome, so breaching it means a units "
                f"slip or a serious mis-estimate.{hint}"
            )
    return out


# ---------------------------------------------------------------------------
# 6. Resolving the number — and never, ever leaving the cell empty
# ---------------------------------------------------------------------------


@dataclass
class Resolved:
    value: float
    provenance: str          # "forecast" | "interval-midpoint" | ... | "zero"
    is_fallback: bool
    notes: list[str] = field(default_factory=list)


def resolve_value(ticker: str, label: str, units: str, metric: SourceMetric | None) -> Resolved:
    """The number that goes in the cell. There is always one.

    A blank cell scores the maximum 5.0 penalty — strictly worse than any number
    a fallback can produce — so the ladder below descends from "the forecast" to
    "the middle of the guardrail band" rather than ever stopping. Every rung
    below the first is flagged as a fallback and shouted about by the caller.
    """
    tk = canonical_ticker(ticker)

    if metric is not None and metric.value is not None:
        return Resolved(float(metric.value), "forecast", False)

    if metric is not None and metric.low is not None and metric.high is not None:
        mid = (metric.low + metric.high) / 2.0
        return Resolved(
            mid,
            "interval-midpoint",
            True,
            [f"point estimate missing; used the midpoint of [{metric.low}, {metric.high}]"],
        )
    if metric is not None and (metric.low is not None or metric.high is not None):
        one = metric.low if metric.low is not None else metric.high
        return Resolved(
            float(one),  # type: ignore[arg-type]
            "interval-endpoint",
            True,
            [f"only one interval endpoint was supplied ({one}); used it as the estimate"],
        )

    bounds = MAGNITUDE_BOUNDS.get((tk, label))
    if bounds:
        lo, hi = bounds
        mid = (lo + hi) / 2.0
        return Resolved(
            mid,
            "sanity-band-midpoint",
            True,
            [
                f"NO forecast and NO interval for this metric. Wrote the midpoint of the "
                f"order-of-magnitude guardrail band [{lo:,g}, {hi:,g}] = {mid:,g}. This is "
                f"NOT a forecast — it is a placeholder that avoids the 5.0 blank penalty."
            ],
        )

    return Resolved(
        0.0,
        "zero",
        True,
        [
            "NO forecast, NO interval and NO sanity band for this metric. Wrote 0.0 purely "
            "so the cell is not empty. This number is meaningless — replace it."
        ],
    )


# ---------------------------------------------------------------------------
# 7. The plan — what would be written, per cell
# ---------------------------------------------------------------------------


@dataclass
class CellPlan:
    ticker: str
    workbook: str
    sheet: str
    cell: str
    label: str
    units: str
    value: float
    provenance: str
    is_fallback: bool
    number_format: str
    notes: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class WorkbookPlan:
    ticker: str
    period: str
    template: Path
    destination: Path
    layout: TemplateLayout
    cells: list[CellPlan] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)
    origin: str = ""

    @property
    def errors(self) -> list[str]:
        return [p for p in self.problems if p.startswith("ERROR")]

    @property
    def fallbacks(self) -> list[CellPlan]:
        return [c for c in self.cells if c.is_fallback]

    @property
    def warnings(self) -> list[str]:
        return [w for c in self.cells for w in c.warnings] + [
            p for p in self.problems if p.startswith("WARN")
        ]


def plan_workbook(ticker: str, source: SourceForecast, out_dir: Path) -> WorkbookPlan:
    """Everything that would happen to one workbook, computed before anything does."""
    tk = canonical_ticker(ticker)
    tpl = template_path(tk)
    name = output_file_for(tk)
    dest = out_dir / name
    problems: list[str] = []

    if not tpl.exists():
        problems.append(
            f"ERROR: template {tpl} is missing. SUBMISSION.md requires the supplied "
            f"template as the starting point; a workbook built from scratch is void."
        )
        return WorkbookPlan(tk, period_for(tk), tpl, dest,
                            TemplateLayout(tk, period_for(tk), SUMMARY_SHEET, 0, 0, ""),
                            [], problems, source.origin)

    layout = read_layout(tpl, tk)
    problems.extend(layout.problems)

    if source.period and source.period != layout.period:
        problems.append(
            f"WARN: {tk}: forecast is labelled period {source.period!r} but the workbook "
            f"is for {layout.period!r}. The workbook header is authoritative and is not "
            f"altered — confirm the forecast targets the right fiscal period."
        )

    cells: list[CellPlan] = []
    for mr in layout.rows:
        metric = _match_metric(source, mr.label)
        if metric is None:
            problems.append(
                f"ERROR: {tk}: the source forecast has no metric matching {mr.label!r} "
                f"(it offered: {[m.label for m in source.metrics] or 'nothing'})."
            )
        elif metric.units and metric.units.strip() != mr.units:
            problems.append(
                f"WARN: {tk}: {mr.label!r} arrived with units {metric.units!r}; the "
                f"workbook cell is {mr.units!r}. The workbook's units are NOT changed — "
                f"CHECK THE SCALE of {metric.value!r}."
            )

        r = resolve_value(tk, mr.label, mr.units, metric)
        cells.append(
            CellPlan(
                ticker=tk,
                workbook=name,
                sheet=layout.sheet,
                cell=mr.forecast_cell,
                label=mr.label,
                units=mr.units,
                value=r.value,
                provenance=r.provenance,
                is_fallback=r.is_fallback,
                number_format=mr.number_format,
                notes=r.notes,
                warnings=unit_warnings(tk, mr.label, mr.units, r.value),
            )
        )

    return WorkbookPlan(tk, layout.period, tpl, dest, layout, cells, problems, source.origin)


def print_plan(plan: WorkbookPlan, *, prefix: str = "would write") -> None:
    say(f"{plan.destination.name}  <- template {plan.template.name}  "
        f"[{plan.ticker} {plan.period}]  source: {plan.origin or 'in-memory'}")
    say(f"    sheet {plan.layout.sheet!r}, header row {plan.layout.header_row}, "
        f"period header {plan.layout.period_header_cell}={plan.period!r} (untouched)")
    for c in plan.cells:
        flag = "  <-- FALLBACK" if c.is_fallback else ""
        say(f"    {prefix} {c.sheet}!{c.cell:<4} = {c.value:>14,.4f}  "
            f"{c.label:<42} [{c.units:<11}] ({c.provenance}){flag}")
        for n in c.notes:
            say(f"        note: {n}")
    for p in plan.problems:
        say(f"    {p}")


# ---------------------------------------------------------------------------
# 8. Writing
# ---------------------------------------------------------------------------


def _write_one(plan: WorkbookPlan) -> Path:
    """Copy the template, assign the three cells, replace the destination atomically.

    The copy-then-fill order is what SUBMISSION.md asks for and it is also the
    only way to be sure the `Instructions` sheet, the styles, the labels and the
    header survive verbatim: nothing here creates a workbook, it only assigns
    three cell values inside one the organisers supplied.

    The temp-file-then-replace keeps a half-written workbook from ever appearing
    at the submission path if this dies mid-save. The temp file keeps an .xlsx
    suffix because openpyxl refuses to open anything else, and it is removed on
    failure so a stale partial cannot be mistaken for a submission.
    """
    dest = plan.destination
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_name(f"~{dest.stem}.partial.xlsx")

    try:
        shutil.copy2(plan.template, tmp)
        wb = openpyxl.load_workbook(tmp)
        try:
            ws = wb[plan.layout.sheet]
            for c in plan.cells:
                cell = ws[c.cell]
                cell.value = float(c.value)  # NUMERIC. A string fails the checker.
            wb.save(tmp)
        finally:
            wb.close()
        os.replace(tmp, dest)
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise
    return dest


def write_workbooks(
    forecasts: Any,
    out_dir: Path | str | None = None,
    *,
    dry_run: bool = False,
    verify: bool = True,
) -> list[Path]:
    """THE ENTRY POINT. Fill the four templates and return the paths written.

    `forecasts` may be anything the rest of the system produces: a list of
    `CompanyForecast`, a single one, a `{ticker: CompanyForecast}` mapping, the
    pipeline's JSON envelopes, or plain dicts. Anything it cannot read becomes a
    loudly-flagged fallback rather than an exception, because the four files
    existing and being numeric matters more than this function's dignity.

    A company with no forecast at all still gets its workbook: the template is
    copied and filled with the fallback ladder in `resolve_value`, and the run
    log carries a banner saying so.
    """
    out = Path(out_dir) if out_dir is not None else DEFAULT_OUT_DIR
    sources = _index_sources(forecasts)

    if _CENTRAL_IMPORT_ERROR:
        shout([
            "analysts.central could not be imported; the metric definitions were",
            f"rebuilt from {COMPANIES_JSON.name} instead. Reason: {_CENTRAL_IMPORT_ERROR}",
            "Labels and units are still authoritative, but validate_forecast did not run.",
        ])

    plans = [plan_workbook(tk, sources.get(tk, SourceForecast(ticker=tk)), out) for tk in TICKERS]

    missing = [tk for tk in TICKERS if tk not in sources]
    if missing:
        shout([
            f"NO FORECAST SUPPLIED FOR: {', '.join(missing)}.",
            "Their workbooks are being filled from the fallback ladder, NOT from a",
            "forecast. Every one of those numbers is a placeholder. Fix the source.",
        ])

    say()
    say("=" * 78)
    say(f"{'DRY RUN — nothing will be written' if dry_run else 'WRITING WORKBOOKS'}")
    say(f"templates : {TEMPLATE_DIR}")
    say(f"output    : {out}")
    say("=" * 78)

    written: list[Path] = []
    for plan in plans:
        say()
        print_plan(plan, prefix="would write" if dry_run else "write")
        if plan.errors:
            shout([f"{plan.destination.name}: {len(plan.errors)} STRUCTURAL ERROR(S)"]
                  + plan.errors)
        if dry_run:
            continue
        if not plan.cells:
            shout([
                f"{plan.destination.name}: NOTHING TO WRITE — the template layout could not",
                "be read, so no cell was located. This workbook will NOT be valid.",
                "Every metric in it scores the maximum 5.0 penalty. Fix the template.",
            ])
            continue
        written.append(_write_one(plan))

    # --- the two things a human must never scroll past ---------------------
    all_fallbacks = [(p, c) for p in plans for c in p.fallbacks]
    if all_fallbacks:
        lines = [f"{len(all_fallbacks)} OF 12 CELLS ARE FALLBACKS, NOT FORECASTS:"]
        for p, c in all_fallbacks:
            lines.append(f"  {c.workbook} {c.cell} {c.label} = {c.value:,g} [{c.provenance}]")
        lines.append("These will score badly. They exist only because a blank scores worse.")
        shout(lines)

    all_warnings = [w for p in plans for w in p.warnings]
    if all_warnings:
        shout(["UNIT / MAGNITUDE WARNINGS — read every line:"] + [f"  {w}" for w in all_warnings])

    if not dry_run:
        say()
        say(f"wrote {len(written)}/{len(TICKERS)} workbooks to {out}")
        for p in written:
            say(f"    {p}")
        if verify:
            verify_workbooks(forecasts=sources, out_dir=out)

    return written


def _index_sources(forecasts: Any) -> dict[str, SourceForecast]:
    """Whatever was handed in -> {canonical ticker: SourceForecast}."""
    if forecasts is None:
        return {}
    if isinstance(forecasts, dict):
        # Either {ticker: forecast} or a single forecast/envelope dict.
        looks_like_map = all(
            isinstance(k, str) and canonical_ticker(k) in TICKERS for k in forecasts
        ) and bool(forecasts)
        if looks_like_map:
            out: dict[str, SourceForecast] = {}
            for k, v in forecasts.items():
                if isinstance(v, SourceForecast):
                    sf = v
                else:
                    sf = normalise(v)
                sf.ticker = sf.ticker or canonical_ticker(k)
                out[canonical_ticker(k)] = sf
            return out
        forecasts = [forecasts]
    elif isinstance(forecasts, (str, Path)):
        return load_forecasts(Path(forecasts))
    elif not isinstance(forecasts, Iterable):
        forecasts = [forecasts]

    out = {}
    for item in forecasts:  # type: ignore[union-attr]
        sf = item if isinstance(item, SourceForecast) else normalise(item)
        tk = canonical_ticker(sf.ticker)
        if tk not in TICKERS:
            shout([
                f"IGNORING a forecast for unknown ticker {sf.ticker!r}.",
                f"challenge/companies.json knows only {', '.join(TICKERS)}.",
            ])
            continue
        out[tk] = sf
    return out


# ---------------------------------------------------------------------------
# 9. Verification — reopen the written files and prove they are right
# ---------------------------------------------------------------------------


def verify_workbooks(
    forecasts: Any = None,
    out_dir: Path | str | None = None,
    *,
    strict: bool = False,
) -> bool:
    """Re-open every written workbook and check it the way the checker will.

    Independent of the write path on purpose: it reads the files from disk,
    re-derives the layout from the label text, and compares the numbers back
    against the source forecast. Passing here is not proof of passing
    check-forecasts.mjs — that script is the arbiter — but everything it tests
    is tested here first, in Python, before npm is involved.
    """
    out = Path(out_dir) if out_dir is not None else DEFAULT_OUT_DIR
    sources = _index_sources(forecasts) if forecasts is not None else {}
    ok = True

    say()
    say("=" * 78)
    say(f"VERIFYING {out}")
    say("=" * 78)

    for tk in TICKERS:
        name = output_file_for(tk)
        path = out / name
        say()
        say(f"{name}  [{tk} {period_for(tk)}]")

        if not path.exists():
            say(f"    FAIL: file is missing at {path}")
            ok = False
            continue

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            say(f"    FAIL: cannot be opened as .xlsx ({e})")
            ok = False
            continue

        try:
            if SUMMARY_SHEET not in wb.sheetnames:
                say(f"    FAIL: no {SUMMARY_SHEET!r} sheet (found {wb.sheetnames})")
                ok = False
                continue
            say(f"    sheets: {wb.sheetnames}")
        finally:
            wb.close()

        layout = read_layout(path, tk)
        for p in layout.problems:
            say(f"    {p}")
            if p.startswith("ERROR"):
                ok = False
        if not layout.header_row:
            ok = False
            continue

        say(f"    header  {layout.sheet}!A{layout.header_row}='Metric' "
            f"B{layout.header_row}='Units' "
            f"{layout.period_header_cell}={period_for(tk)!r}")
        say(f"    {'cell':<5} {'metric':<42} {'units':<11} {'written':>16} "
            f"{'source':>16}  {'type':<7} status")
        say(f"    {'-' * 5} {'-' * 42} {'-' * 11} {'-' * 16} {'-' * 16}  {'-' * 7} ------")

        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = wb[layout.sheet]
            src = sources.get(tk)
            for mr in layout.rows:
                cell = ws[mr.forecast_cell]
                v = cell.value
                status = "OK"
                if v is None:
                    status = "FAIL empty cell"
                    ok = False
                elif not isinstance(v, (int, float)) or isinstance(v, bool):
                    status = f"FAIL not numeric ({type(v).__name__})"
                    ok = False
                elif not math.isfinite(float(v)):
                    status = "FAIL not finite"
                    ok = False

                expected_txt = ""
                if src is not None:
                    sm = _match_metric(src, mr.label)
                    if sm is not None and sm.value is not None:
                        expected_txt = f"{sm.value:,.4f}"
                        if isinstance(v, (int, float)) and not math.isclose(
                            float(v), float(sm.value), rel_tol=1e-9, abs_tol=1e-9
                        ):
                            status = f"FAIL value {v} != forecast {sm.value}"
                            ok = False
                    else:
                        expected_txt = "(fallback)"

                written_txt = f"{v:,.4f}" if isinstance(v, (int, float)) else repr(v)
                say(f"    {mr.forecast_cell:<5} {mr.label:<42} {mr.units:<11} "
                    f"{written_txt:>16} {expected_txt:>16}  "
                    f"{type(v).__name__:<7} {status}")

                for w in unit_warnings(tk, mr.label, mr.units, float(v)) if isinstance(
                    v, (int, float)
                ) else []:
                    say(f"          WARNING {w}")
                    if strict:
                        ok = False
        finally:
            wb.close()

        # The pydantic-level validator, when the agent stack is importable.
        if validate_forecast is not None and src is not None and src.raw is not None:
            try:
                from analysts.models import CompanyForecast  # local: optional dep

                cf = src.raw if isinstance(src.raw, CompanyForecast) else None
                if cf is not None:
                    for p in validate_forecast(tk, cf):
                        say(f"    central.validate_forecast: {p}")
            except Exception:
                pass

    say()
    say("=" * 78)
    say("VERIFY PASSED — all four workbooks have a Summary sheet, the expected "
        "labels and units, and numeric forecasts." if ok else
        "VERIFY FAILED — see the FAIL lines above.")
    say("=" * 78)
    if not ok:
        shout([
            "THE SUBMISSION IS NOT VALID. Do not upload. Fix the failures above and",
            "re-run, then confirm with `npm run check:forecasts` from inside starter/.",
        ])
    return ok


# ---------------------------------------------------------------------------
# 10. Reading the pipeline's forecast JSON off disk
# ---------------------------------------------------------------------------


def load_forecasts(source: Path | str) -> dict[str, SourceForecast]:
    """Read forecasts from the directory (or file) the pipeline wrote.

    `analysts.pipeline.write_outputs` writes one `<workbook stem>.json` per
    company — an envelope around the `CompanyForecast` — plus a combined
    `_forecasts.json`. Both are accepted, along with a bare `CompanyForecast`
    dump, because the input format is not worth failing the submission over.
    """
    p = Path(source)
    if not p.is_absolute() and not p.exists():
        alt = REPO_ROOT / p
        if alt.exists():
            p = alt

    out: dict[str, SourceForecast] = {}
    if not p.exists():
        shout([
            f"FORECAST SOURCE NOT FOUND: {p}",
            "No numbers to write. Every cell would be a fallback placeholder.",
        ])
        return out

    if p.is_file():
        for tk, sf in _read_forecast_file(p).items():
            out[tk] = sf
        return out

    # Per-company files first — they carry the full CompanyForecast.
    for tk in TICKERS:
        stem = (output_file_for(tk) or f"{tk}.xlsx").rsplit(".", 1)[0]
        f = p / f"{stem}.json"
        if f.exists():
            found = _read_forecast_file(f)
            if tk in found:
                out[tk] = found[tk]

    # Then the combined file, for anything still missing.
    combined = p / "_forecasts.json"
    if combined.exists():
        for tk, sf in _read_forecast_file(combined).items():
            out.setdefault(tk, sf)

    # Last resort: anything else in the directory.
    if len(out) < len(TICKERS):
        for f in sorted(p.glob("*.json")):
            for tk, sf in _read_forecast_file(f).items():
                out.setdefault(tk, sf)

    return out


def _read_forecast_file(path: Path) -> dict[str, SourceForecast]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        shout([f"UNREADABLE forecast file {path}: {type(e).__name__}: {e}"])
        return {}

    out: dict[str, SourceForecast] = {}

    def take(obj: Any) -> None:
        sf = normalise(obj, origin=str(path))
        tk = canonical_ticker(sf.ticker)
        if tk in TICKERS:
            out.setdefault(tk, sf)

    if isinstance(data, dict) and isinstance(data.get("companies"), list):
        for entry in data["companies"]:
            take(entry)
    elif isinstance(data, list):
        for entry in data:
            take(entry)
    elif isinstance(data, dict):
        take(data)
    return out


# ---------------------------------------------------------------------------
# 11. Synthetic forecasts — for exercising the writer without the agent stack
# ---------------------------------------------------------------------------

# Plausible, inside every sanity band, and deliberately NOT anyone's forecast.
# Used by `--synthetic` to prove the writer end to end (including the checker)
# without spending a model call or touching a live run's artefacts.
_SYNTHETIC: dict[str, list[float]] = {
    "HD": [45_500.0, 4.72, 1.5],
    "ADI": [2_950.0, 2.05, 69.5],
    "HAS": [1_015.0, 3.9, 55.0],
    "DE": [11_200.0, 4.35, 1_150.0],
}


def synthetic_forecasts() -> dict[str, SourceForecast]:
    out: dict[str, SourceForecast] = {}
    for tk in TICKERS:
        values = _SYNTHETIC.get(tk, [])
        metrics = []
        for i, spec in enumerate(metrics_for(tk)):
            v = values[i] if i < len(values) else None
            metrics.append(
                SourceMetric(
                    label=spec["label"],
                    units=spec["units"],
                    value=v,
                    low=None if v is None else v * 0.97,
                    high=None if v is None else v * 1.03,
                )
            )
        out[tk] = SourceForecast(
            ticker=tk, period=period_for(tk), metrics=metrics, origin="synthetic"
        )
    return out


# ---------------------------------------------------------------------------
# 12. CLI
# ---------------------------------------------------------------------------


def cli(argv: Sequence[str] | None = None) -> int:
    _stdout_utf8()
    ap = argparse.ArgumentParser(
        prog="python -m submission.workbook",
        description="Fill the four supplied .xlsx templates with the forecasts and "
                    "write them to the submission folder.",
    )
    ap.add_argument("--from", dest="source", default=str(DEFAULT_FORECAST_DIR),
                    help="Directory (or file) of forecast JSON written by the pipeline. "
                         f"Default: {DEFAULT_FORECAST_DIR}")
    ap.add_argument("--out", default=str(DEFAULT_OUT_DIR),
                    help=f"Where the four workbooks go. Default: {DEFAULT_OUT_DIR}")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print every cell that would be written; write nothing.")
    ap.add_argument("--synthetic", action="store_true",
                    help="Ignore --from and use built-in plausible test values.")
    ap.add_argument("--verify-only", action="store_true",
                    help="Do not write; re-check the workbooks already in --out.")
    ap.add_argument("--dump-templates", action="store_true",
                    help="Print the discovered structure of the four templates and exit.")
    ap.add_argument("--no-verify", action="store_true",
                    help="Skip the post-write verification pass.")
    ap.add_argument("--strict", action="store_true",
                    help="Treat unit/magnitude warnings as failures.")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )

    if args.dump_templates:
        dump_templates()
        return 0

    out_dir = Path(args.out)
    if not out_dir.is_absolute():
        out_dir = (Path.cwd() / out_dir) if (Path.cwd() / out_dir).parent.exists() \
            else (REPO_ROOT / out_dir)

    if args.verify_only:
        sources = synthetic_forecasts() if args.synthetic else load_forecasts(args.source)
        return 0 if verify_workbooks(sources, out_dir, strict=args.strict) else 1

    if args.synthetic:
        sources: Any = synthetic_forecasts()
        say("using SYNTHETIC forecasts — these are test values, not a submission.")
    else:
        sources = load_forecasts(args.source)
        say(f"loaded {len(sources)}/{len(TICKERS)} forecasts from {args.source}")

    written = write_workbooks(
        sources, out_dir, dry_run=args.dry_run, verify=not args.no_verify
    )

    if args.dry_run:
        return 0
    if len(written) != len(TICKERS):
        shout([
            f"ONLY {len(written)} OF {len(TICKERS)} WORKBOOKS WERE WRITTEN.",
            "The submission is incomplete. Every missing file is three metrics at the",
            "maximum 5.0 penalty.",
        ])
        return 1
    if args.no_verify:
        return 0
    return 0 if verify_workbooks(sources, out_dir, strict=args.strict) else 1


__all__ = [
    "CellPlan",
    "SourceForecast",
    "SourceMetric",
    "TemplateLayout",
    "WorkbookPlan",
    "dump_templates",
    "load_forecasts",
    "normalise",
    "plan_workbook",
    "read_layout",
    "resolve_value",
    "synthetic_forecasts",
    "template_path",
    "unit_warnings",
    "verify_workbooks",
    "write_workbooks",
]


if __name__ == "__main__":
    raise SystemExit(cli())
