"""Workbook writer.

Loads the organizer template and writes only the forecast cells. Never constructs a
workbook from scratch - the Summary sheet structure, metric labels, units and period
header are the submission contract, and `check-forecasts.mjs` fails the entry if any of
them drift.

Contract (verified against scripts/check-forecasts.mjs and the templates):
  - sheet named "Summary"
  - a header row within rows 1-30 where A="Metric", B="Units", C=<period>
  - the next N rows: A=exact metric label, B=exact units, C=a finite number
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_DIR = Path("challenge/templates")
DEFINITIONS = Path("challenge/companies.json")


class WorkbookContractError(RuntimeError):
    """The template does not look like we expect. Fail loudly rather than upload junk."""


def _find_header_row(ws, period: str) -> int:
    for row in range(1, 31):
        if (
            str(ws.cell(row=row, column=1).value or "").strip() == "Metric"
            and str(ws.cell(row=row, column=2).value or "").strip() == "Units"
            and str(ws.cell(row=row, column=3).value or "").strip() == period
        ):
            return row
    raise WorkbookContractError(
        f"No 'Metric | Units | {period}' header found in rows 1-30"
    )


def write_workbook(
    ticker_file: str,
    period: str,
    metrics: list[dict],
    values: dict[str, float],
    out_dir: str | Path = "submission",
    template_dir: str | Path = TEMPLATE_DIR,
) -> Path:
    """Write one company's workbook.

    `metrics` is the ordered metric list from companies.json.
    `values` maps metric label -> number. A missing or non-finite value raises, because a
    blank cell scores 5.0 and we would rather fail loudly at build time than silently
    upload an empty forecast.
    """
    template = Path(template_dir) / ticker_file
    if not template.exists():
        raise FileNotFoundError(f"Template missing: {template}")

    wb = load_workbook(template)
    if "Summary" not in wb.sheetnames:
        raise WorkbookContractError(f"{ticker_file}: no Summary sheet")
    ws = wb["Summary"]

    header = _find_header_row(ws, period)

    for offset, metric in enumerate(metrics):
        row = header + offset + 1
        label_cell = str(ws.cell(row=row, column=1).value or "").strip()
        units_cell = str(ws.cell(row=row, column=2).value or "").strip()

        if label_cell != metric["label"]:
            raise WorkbookContractError(
                f"{ticker_file} row {row}: expected label {metric['label']!r}, "
                f"template has {label_cell!r}"
            )
        if units_cell != metric["units"]:
            raise WorkbookContractError(
                f"{ticker_file} row {row}: expected units {metric['units']!r}, "
                f"template has {units_cell!r}"
            )

        value = values.get(metric["label"])
        if value is None or not isinstance(value, (int, float)) or value != value:
            raise ValueError(
                f"{ticker_file}: no finite value for {metric['label']!r}. "
                "A blank cell scores 5.0 - emit the anchor instead of nothing."
            )
        ws.cell(row=row, column=3).value = float(value)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / ticker_file
    wb.save(out_path)
    return out_path


def write_all(
    forecasts: dict[str, dict[str, float]],
    definitions_path: str | Path = DEFINITIONS,
    out_dir: str | Path = "submission",
    template_dir: str | Path = TEMPLATE_DIR,
) -> list[Path]:
    """Write all four workbooks. `forecasts` maps ticker -> {metric label: value}."""
    definitions = json.loads(Path(definitions_path).read_text(encoding="utf-8"))
    written = []
    for company in definitions["companies"]:
        values = forecasts.get(company["ticker"], {})
        written.append(
            write_workbook(
                ticker_file=company["outputFile"],
                period=company["period"],
                metrics=company["metrics"],
                values=values,
                out_dir=out_dir,
                template_dir=template_dir,
            )
        )
    return written
